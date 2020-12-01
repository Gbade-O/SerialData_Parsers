# -*- coding: utf-8 -*-
"""
Created on Wed Nov 18 15:36:35 2020

@author: GOgunwumi
"""

import os 
import pandas as pd 
import matplotlib.pyplot as plt
import numpy as np
import collections as col
import xlsxwriter
from openpyxl import load_workbook
import csv 
import re 
import codecs
import glob
from openpyxl.styles import numbers




#Function takes in a title and creates worksheet according to title & headers 
def CreateSheet(title,Header): 
    worksheet = workbook.add_worksheet(title) #Create empty worksheet 
    #Create specific formats 
    header_format = workbook.add_format({'bold': True, 'valign' : 'vjustify'})
    
    
    #Write header to worksheet 
    for item in enumerate(Header):
            
        worksheet.write(0,item[0],item[1],header_format)
        
    return worksheet
    
    


#Create Workboook 
workbook = xlsxwriter.Workbook("CompiledData.xlsx")
Headers  = ["Date","Brew Number","Brew Size","PreTest Resevoir Weight (g)","Pretest Carafe Weight (g)","Post Test Resevoir Weight (g)","Post Test Carafe Weight (g)","Brew Time", "Comments/Notes","Actual Water Used (g)", "Volume output (g)", "Vol. Error"]


#define some formats 
brewT_format = workbook.add_format({'num_format': 'h:mm;@'})
VolErr_format = workbook.add_format({'num_format': '0.00%'})
Comments_format = workbook.add_format({'num_format': 'General', 'text_wrap': True })
Data_format = workbook.add_format({'num_format': 'General','align': 'center'})


BrewFilePath = r"C:\Users\gogunwumi\Documents\Temp\BrewData"

os.chdir(BrewFilePath) ##Change to directory for Brewfiles 

files = [ name for name in os.listdir(".") if os.path.isfile(name)]

for file in files :
    
    #read data into Pandas 
    xls = pd.ExcelFile(file)
    sheets = xls.sheet_names #Document will have multiple sheets to read from 
    
    
    #Loop through sheets , copy data to new sheet 
    for name in sheets :
        
        #Get SKU
        SKU = name.split('_')[0]
        #create new sheet 
        worksheet = CreateSheet(name, Headers)
        worksheet.set_column(8,8,30) #Set the comments column width to be wider 
        
        #import data from source spreasheet 
        DataToSend = xls.parse(name)
        
        #Double for loop to write columns and rows 
        nrows  = len(DataToSend)
        columns = DataToSend.columns 
        
        #Actual number of rows counter 
        row_cnt = 0
        
        #Write to new sheet , cell by cell 
        for i in range(1,nrows):
            
            for column in enumerate(columns) :
                
                info = DataToSend.iloc[i-1,column[0]] #Get value of information to write 
                
                try :
                    #These columns can be treated the same 
                    if column[0] > 2 and column[0] < 7 :
                        
                        worksheet.write(i,column[0],info,Data_format)
                    
                    if column[1] == "Date":
                        
                        print(info)
                        
                    if "Brew Time" in column[1]:
                        
                        worksheet.write(i,column[0],info,brewT_format)
                         
                    if "Comments/Notes" in column[1] :
                        
                        worksheet.write(i,column[0],info,Comments_format)
                    
                    if "Brew Size" in column[1]:
                        
                        Brew_Target = info
                        
                    
                                   
                except :
                    
                    pass
                
            #Perform some calculations, now that most data in row has been written
            
            #Calculate Actual Water Used 
            worksheet.write_formula(i,9,'{=D' + str(i+1) + '- F'+str(i+1)+'}')
            
            #Calculate Volume Output 
            worksheet.write_formula(i,10,'{=G' + str(i+1) + '- E'+str(i+1)+'}')
            
            #Calculate Vol Error
            
            worksheet.write_formula(i,11,'{=(J' + str(i+1) + '-'+str(Brew_Target)+')/'+ str(Brew_Target)+'}',VolErr_format)
    
                    
workbook.close()
                    
                        
                
            
            
        
        
        
        
        
        
        
    












# MasterFilePath = r"C:\Users\gogunwumi\Documents\Temp\Scale Testing CFP300.xlsx"

# ##Create Writer to MainSheet 
# writer = pd.ExcelWriter(MasterFilePath, engine = 'openpyxl')
   
# ##read in workbook
# writer.book = load_workbook(MasterFilePath)

# BrewFilePath = r"C:\Users\gogunwumi\Documents\Temp\BrewData"

# os.chdir(BrewFilePath) ##Change to directory for Brewfiles 

# files = [ name for name in os.listdir(".") if os.path.isfile(name)]


# for file in files :
    
#     xls = pd.ExcelFile(file)
#     sheets = xls.sheet_names
    
#     for name in sheets:
        
#         df = pd.read_excel(file, sheet_name = name)
        
       
        
#         ##Get/Find appropriate Sheet 
#         Sheet = writer.book[name]
        
#         #Set revelant formats before hand 
        
        
#         #Brute force copy data to destination sheet 
        
#         for i in range(2, len(df)):
            
           
#             #Write nearby columns in for loop 
#             for column in enumerate(Sheet.iter_cols(min_row = 1, max_row = 1, min_col = 4, max_col =9 , values_only = True),start = 4):
                
#                 Sheet.cell(row = i, column = column[0]).value = df.iloc[i-2,column[0]-1]
                
           
               
            
       
            

# writer.book.save(MasterFilePath)

        
        
        
        
        
        
 
    
    